"""Reference solver: every line item still open in the firm's working papers.

Every other task in this suite is answered by querying a surface. This one
is not answerable that way at all. The working papers are eighty-five real
workbooks in the firm's shared folders — two hundred and thirty-eight
sheets, fourteen thousand cells — and no MCP server serves them, no SQL
table holds them, and nothing indexes them. The only route is to open
every file.

That is the point of it. Everything measured on this world so far has been
solved by paginating an API once, writing the result to disk, and running
code over the copy: volume, chained derivation, lexical near-misses and
semantic synonyms have each been tried and each came back at ceiling. A
library of spreadsheets is a different shape of work — the enumeration is
the task, and two workbooks in three contain nothing, so sampling looks
like it is working right up until the totals come out wrong.

The rule itself is deliberately unambiguous: a status is open when the
cell equals one of five strings, case-folded and trimmed, and not when it
merely contains one. The firm writes `Pending PBC`, `Open — final figure
confirmed` and `Not started — count 12/31 done, sheets pending`, and a
substring match takes in thirty-odd rows the rule excludes. The
instruction lists those spellings outright rather than leaving them to be
discovered, because a rule an agent has to infer is a task defect
whichever way the agent guesses.
"""

import datetime
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

WORKSPACE = Path(os.environ.get("WORKBENCH_WORKSPACE", "."))
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("open_items.json")

# Case-folded, because the working papers carry both `Not started` and
# `Not Started` and they are the same item in anybody's reading.
OPEN = frozenset({"open", "outstanding", "pending", "not started", "blocked"})

# The library is not perfectly consistent, and one due date is written
# `1/15/2025` while a hundred and thirteen are ISO. That single cell is
# also the earliest date in the firm, so `earliest_due_date` is exactly
# the figure that separates normalising the values from sorting them as
# strings -- which is why the instruction says outright that the wrinkle
# is there rather than leaving it to be tripped over.
_US = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")


def _as_date(value: str) -> datetime.date | None:
    try:
        return datetime.date.fromisoformat(value)
    except ValueError:
        pass
    match = _US.fullmatch(value)
    if match:
        month, day, year = (int(part) for part in match.groups())
        return datetime.date(year, month, day)
    return None


def _cell(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def main() -> None:
    rows: list[dict] = []
    workbooks = sheets = 0

    # Sorted, so the answer does not depend on the order the filesystem
    # happens to hand back.
    for path in sorted(WORKSPACE.rglob("*.xlsx")):
        workbooks += 1
        book = load_workbook(path, data_only=True)
        relative = path.relative_to(WORKSPACE).as_posix()
        for sheet in book.worksheets:
            sheets += 1
            header = [
                str(cell).strip() if cell is not None else ""
                for cell in next(sheet.iter_rows(max_row=1, values_only=True), ())
            ]
            if "Status" not in header:
                continue
            status_at = header.index("Status")
            owner_at = header.index("Owner") if "Owner" in header else None
            due_at = header.index("Due Date") if "Due Date" in header else None
            # Row 1 is the header, so the first item is row 2 -- the number
            # a person reads off the side of the spreadsheet.
            for number, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                status = _cell(row, status_at)
                if status.casefold() not in OPEN:
                    continue
                rows.append(
                    {
                        "workbook": relative,
                        "sheet": sheet.title,
                        "row": number,
                        "status": status,
                        "owner": _cell(row, owner_at),
                        "due_date": _cell(row, due_at),
                    }
                )
        book.close()

    rows.sort(key=lambda r: (r["workbook"], r["sheet"], r["row"]))
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    dues = sorted(
        date for date in (_as_date(r["due_date"]) for r in rows if r["due_date"])
        if date is not None
    )

    OUT.write_text(
        json.dumps(
            {
                "workbooks_read": workbooks,
                "sheets_read": sheets,
                "open_items_total": len(rows),
                "workbooks_with_open_items": len({r["workbook"] for r in rows}),
                # Most, then the earlier spelling -- `max` breaks a tie the
                # other way and the instruction says alphabetically first.
                "top_status": min(counts, key=lambda s: (-counts[s], s))
                if counts
                else "",
                "earliest_due_date": dues[0].isoformat() if dues else "",
                "open_items": rows,
            },
            indent=1,
        )
        + "\n"
    )


if __name__ == "__main__":
    main()
