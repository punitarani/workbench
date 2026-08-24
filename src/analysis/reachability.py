"""What the tool surfaces actually serve, and what an oracle demands of them.

An oracle extracted from world state is not automatically answerable. The
reference solvers open ``state/*.db`` directly, so a task rule can name a
column no tool returns — and the resulting score measures whether the
agent guessed an internal vocabulary, not whether it did the work.

That is not hypothetical. ``tkt-000005`` is the world's own ticket id; the
clio surface identifies a matter by ``id`` and ``display_number``
(``00005-Mensah``) and never emits the ``tkt-`` form, which leaked instead
into 26 of 65 Slack messages. Two rollouts of one model on one task scored
1.000 and 0.273 on which vocabulary each happened to find.

So: collect every string the servers hand back through their discovery
tools, and require an oracle's identifiers to appear in that set. A task
whose answer key cannot be spelled in the language of its own tools is a
broken task, and this is the gate that says so before it is ever run.
"""

import asyncio
import hashlib
import json
import re
import shutil
import tempfile
from pathlib import Path
from typing import Any

from tools.framework import build_server
from tools.registry import REGISTRY

# Discovery tools take no required argument: they are how an agent with no
# prior knowledge finds anything at all. If a value cannot be reached from
# them (directly or through the ids they hand out), it is not discoverable.
# Sized to the world, not to a round number. A six-month Merrick record
# serves 658 documents, and a non-head version id appears only in the
# result of `get_document_versions(<that document>)` -- so a cap below the
# document count silently decides that some documents' versions are
# unreachable, and the gate then blocks a task whose key is correct. At
# 400 the crawl opened 221 of 658 and reported eight true version ids as
# unservable.
#
# This is a performance bound and never a semantic one. Whenever it binds,
# the honest reading is "the crawl ran out of budget", not "no agent could
# reach this" -- so it is set above the largest surface the datasets carry
# rather than tuned down to what is fast.
_MAX_FOLLOW = 400

# How many times the crawl opens what the previous round named. Two was
# one short: iManage needs workspace -> document -> versions before a
# non-head version id appears anywhere.
_MAX_DEPTH = 3

# Pagination is navigation, not a special case: a message on page nine is
# every bit as reachable as one on page one.
_MAX_PAGES = 60

# Read verbs only. The write half of these surfaces is real — `create_draft`
# takes no required argument and will happily file a draft — and a gate that
# measures a world must not change it.
_READ_VERBS = ("get_", "list_", "search_", "read_", "find_", "who_", "describe_")


def _is_read(name: str) -> bool:
    """A read verb, with or without its system's prefix.

    Slack names every tool ``slack_search_channels``, ``slack_read_thread``
    and so on, so matching the verb at the start of the string excluded that
    whole surface from the crawl -- and the gate then reported 272 message
    timestamps as unservable when the tools serve them plainly.
    """

    return name.startswith(_READ_VERBS) or name.split("_", 1)[-1].startswith(
        _READ_VERBS
    )


# An ISO calendar date, with or without a time on it.
_DATE = re.compile(r"\d{4}-\d{2}-\d{2}(?:[T ][\d:.+\-]*)?")


def _strings(value: Any, into: set[str]) -> None:
    if isinstance(value, str):
        into.add(value)
    elif isinstance(value, dict):
        for key, item in value.items():
            into.add(key)
            _strings(item, into)
    elif isinstance(value, (list, tuple)):
        for item in value:
            _strings(item, into)
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        into.add(str(value))


async def _call(server, name: str, arguments: dict) -> Any:
    try:
        result = await server.call_tool(name, arguments)
    except Exception:  # a tool needing a seat or a different id shape
        return None
    if result.is_error:
        return None
    for chunk in result.content:
        text = getattr(chunk, "text", None)
        if text is None:
            continue
        try:
            return json.loads(text)
        except ValueError:
            return text
    return None


async def _collect(server, name: str, arguments: dict, into: set[str]) -> None:
    """Call a tool and read every page of it.

    Reachability is not "appears on page one". A message on the ninth page
    of a search is still something the agent can find, and a gate that
    stopped at the first page would condemn perfectly good tasks.
    """

    properties = set()
    for tool in await server.list_tools():
        if tool.name == name:
            properties = set((tool.input_schema or {}).get("properties") or {})
            break
    token_arg = next(
        (p for p in ("pageToken", "cursor", "page") if p in properties), None
    )

    arguments = dict(arguments)
    for _ in range(_MAX_PAGES):
        payload = await _call(server, name, arguments)
        if payload is None:
            return
        _strings(payload, into)
        if token_arg is None or not isinstance(payload, dict):
            return
        # Slack tucks its cursor inside `response_metadata`, so looking only
        # at the top level stopped every channel at page one and left most
        # of the firm's chat unreached — which the gate then reported as
        # unservable rather than as its own blind spot.
        nested = payload.get("response_metadata")
        places = (payload, nested if isinstance(nested, dict) else {})
        # `next_page` is iManage's: `get_container_children` returns it and
        # no `hasMore`, so the `page` branch below saw an absent `hasMore`,
        # concluded there was nothing more, and stopped every container at
        # its first hundred children. On a 658-document world that left 437
        # documents unopened and their version ids reported unservable --
        # the gate blaming a task for its own blind spot, again.
        token = next(
            (
                where[key]
                for where in places
                for key in ("nextPageToken", "nextCursor", "next_cursor", "next_page")
                if where.get(key)
            ),
            None,
        )
        if token_arg == "page":
            # Only invent the next page number when the payload said
            # nothing about one. A surface that names its own next page --
            # iManage returns `next_page` and no `hasMore` -- is believed,
            # and a surface that says `hasMore: false` is finished.
            if token is None:
                if not payload.get("hasMore") and not payload.get("has_more"):
                    return
                token = (arguments.get("page") or 1) + 1
        if not token:
            return
        arguments[token_arg] = token


async def _follow(server, tools, discovered: set[str]) -> set[str]:
    """Open what the previous round named, until nothing new turns up.

    This was a single round, and one round short of the surface it gates.
    On iManage an agent goes workspace -> document -> versions:
    `get_container_children` needs a workspace id, and only its result
    names the documents whose version lists carry every id but the head.
    Following step one alone, a six-month world's crawl discovered 105
    `LEGAL!` ids and not one bare document reference, so `LEGAL!24.6` --
    which `get_document_versions("24")` returns without complaint -- was
    reported unservable and blocked a task whose answer key was correct.

    A false positive here is not a small thing. The gate's whole job is to
    say "no agent could produce this", and when it says so wrongly, the
    obvious response is to rewrite a rule that was right.

    Shortest first, then alphabetical. The cap bounds a walk that is
    otherwise quadratic, but truncating it alphabetically sorted
    seventeen-character timestamps ahead of nine-character channel ids and
    cut every channel out of the follow -- so two messages in a
    543-message room were reported as unservable while the tool returned
    them happily when asked.

    Each round follows only what it has not followed already, so the cost
    is rounds x cap rather than the square of everything seen.
    """

    followers = [
        tool
        for tool in tools
        if len((tool.input_schema or {}).get("required") or []) == 1
    ]
    reached: set[str] = set()
    frontier = discovered
    followed: set[str] = set()
    for _ in range(_MAX_DEPTH):
        candidates = sorted(
            (
                value
                for value in frontier
                if value
                and len(value) < 64
                and " " not in value
                and value not in followed
            ),
            key=lambda value: (len(value), value),
        )[:_MAX_FOLLOW]
        if not candidates:
            break
        followed.update(candidates)
        found: set[str] = set()
        for tool in followers:
            required = (tool.input_schema or {}).get("required") or []
            for value in candidates:
                await _collect(server, tool.name, {required[0]: value}, found)
        reached |= found
        frontier = found
    return reached


async def _served(state_dir: Path) -> set[str]:
    """Everything two steps of honest navigation can reach.

    An agent starts with no ids: it lists and searches, then opens what
    those results named. Gmail is the reason the second step matters —
    `search_threads` returns threads, and a message id only appears once
    you open one.
    """

    reachable: set[str] = set()
    for system in REGISTRY:
        db_path = state_dir / f"{system.name}.db"
        if not db_path.is_file():
            continue
        server = build_server(system, db_path)
        tools = [tool for tool in await server.list_tools() if _is_read(tool.name)]
        discovered: set[str] = set()
        for tool in tools:
            required = (tool.input_schema or {}).get("required") or []
            if not required:
                await _collect(server, tool.name, {}, discovered)
            elif required == ["query"]:
                # A search that insists on a query is still discovery: an
                # agent with no ids types an empty one, and so does this.
                # Slack is the case that matters -- `search_channels`
                # requires a query, so skipping it left the whole surface
                # unreached, and 272 message timestamps the tools plainly
                # do serve were reported as unservable.
                await _collect(server, tool.name, {"query": ""}, discovered)
        reachable |= discovered

        followed = await _follow(server, tools, discovered)
        reachable |= followed
        reachable |= await _enumerate(server, tools, discovered | followed)
    return reachable


# Tools that list the contents of one thing, rather than answering a
# question about it. Their results are an enumeration: an agent that holds
# the container holds every child, in one call, with no searching.
_ENUMERATORS = ("get_container_children", "get_document_versions")


async def _enumerate(server, tools, known: set[str]) -> set[str]:
    """Exhaust the enumerating tools over everything already discovered.

    `_follow` samples. It sorts the frontier, takes the first
    `_MAX_FOLLOW`, and calls every single-argument tool on each -- which is
    quadratic, costs eight and a half minutes at four hundred, and on a
    658-document world opened 221 of them. The 437 it did not open had
    every one of their version ids reported unservable, and the gate then
    blocked a task whose answer key was correct and whose rows an agent
    could have obtained in one call each.

    Sampling is the wrong shape for an enumeration. `get_container_children`
    on nine workspaces names every document in the library, and
    `get_document_versions` on a document names every version of it: there
    is nothing to prioritise, because the answer is the whole list. So
    these two run over everything the crawl has seen, and they run to
    completion.

    Linear, and cheap for it: nine containers and six hundred documents is
    two calls per document, against the twenty-odd tools `_follow` would
    have tried on each. Two rounds, because the documents that make a
    version list are named by the container listing.
    """

    enumerators = [tool for tool in tools if tool.name in _ENUMERATORS]
    if not enumerators:
        return set()
    reached: set[str] = set()
    frontier = known
    for _ in range(2):
        found: set[str] = set()
        for tool in enumerators:
            required = (tool.input_schema or {}).get("required") or []
            if len(required) != 1:
                continue
            for value in frontier:
                if not value or " " in value or len(value) >= 64:
                    continue
                await _collect(server, tool.name, {required[0]: value}, found)
        new_values = found - reached - known
        reached |= found
        if not new_values:
            break
        frontier = new_values
    return reached


def served_vocabulary(state_dir: Path) -> set[str]:
    """Every string an agent can obtain from the tools without knowing an id.

    Cached per state directory and per state mtime. The crawl follows four
    hundred candidates through every read tool on five surfaces, sixty pages
    deep, and a build asks the same question once per task — ten identical
    crawls of one unchanging bundle. The mtime is in the key so a rebuilt
    bundle is never answered from a stale one.
    """

    fingerprint = _state_digest(state_dir)
    if fingerprint is None:
        # Unknown provenance: crawl, and cache nothing. This is the path a
        # bundle with no SOURCE takes, and it must not share a key with
        # any other bundle.
        with tempfile.TemporaryDirectory() as scratch:
            mirror = Path(scratch) / "state"
            shutil.copytree(state_dir, mirror)
            return asyncio.run(_served(mirror))
    if fingerprint not in _VOCABULARY:
        _VOCABULARY.clear()
        if (cached := _read_cache(state_dir, fingerprint)) is not None:
            _VOCABULARY[fingerprint] = cached
            return cached
        # Crawled on a copy, because reading this world writes to it.
        # iManage keeps an access log and every tool call appends a row to
        # it -- a real feature of the product, and the reason the staged
        # bundle carried 2,280 file accesses that no person at the firm
        # ever made. The gate's own contract is that measuring a world must
        # not change it, and on the original it was breaking that contract
        # and defeating this cache in the same stroke.
        with tempfile.TemporaryDirectory() as scratch:
            mirror = Path(scratch) / "state"
            shutil.copytree(state_dir, mirror)
            _VOCABULARY[fingerprint] = asyncio.run(_served(mirror))
        _write_cache(state_dir, fingerprint, _VOCABULARY[fingerprint])
    return _VOCABULARY[fingerprint]


# The crawl is the dominant cost of a build: ~18 minutes of the ~25 a
# single-task build takes, and it is repeated in full every time because
# `materialize` rewrites `state/` from scratch and the in-process memo
# dies with the process. Keyed on the *content* of the databases rather
# than their mtimes, so a bundle rebuilt from the same world log answers
# from the cache and a bundle rebuilt from a different one cannot.
#
# Hashing 16MB of SQLite costs well under a second. Getting this wrong in
# the stale direction would be serious -- a reachability verdict from
# another world -- which is why it is a digest of the bytes and not a
# timestamp.
_CACHE_NAME = ".reachability.json"


def _state_digest(state_dir: Path) -> str | None:
    """What this crawl's answer actually depends on: the world log.

    Hashing the databases was the obvious thing and it does not work. Two
    SQLite files holding identical rows differ byte for byte after a
    rebuild — page layout, freelists, insertion order — so the digest
    changed every time `materialize` ran, which is exactly and only when
    the cache would have paid. A test caught it; the implementation looked
    right.

    The served state is a deterministic projection of one world log, so
    the log is the honest key. `SOURCE` records which log built this
    bundle — written by the build the moment the bundle becomes that
    world — and hashing 47MB of JSONL costs about a fifth of a second
    against the eighteen minutes it saves.

    Returns None when the provenance is unknown, which means crawl. A
    missing key must never be treated as a matching one.
    """

    source = state_dir.parent / "SOURCE"
    if not source.is_file():
        return None
    log = Path(source.read_text(encoding="utf-8").strip())
    if not log.is_file():
        return None
    digest = hashlib.sha256()
    with log.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def _read_cache(state_dir: Path, fingerprint: str) -> set[str] | None:
    path = state_dir.parent / _CACHE_NAME
    if not path.is_file():
        return None
    try:
        stored = json.loads(path.read_text(encoding="utf-8"))
    except ValueError:
        return None
    if stored.get("state") != fingerprint:
        return None
    values = stored.get("reachable")
    return set(values) if isinstance(values, list) else None


def _write_cache(state_dir: Path, fingerprint: str, reachable: set[str]) -> None:
    path = state_dir.parent / _CACHE_NAME
    try:
        path.write_text(
            json.dumps({"state": fingerprint, "reachable": sorted(reachable)}),
            encoding="utf-8",
        )
    except OSError:
        # A cache that cannot be written is a slow build, not a wrong one.
        pass


_VOCABULARY: dict[tuple, set[str]] = {}


def _identifiers(oracle: Any, into: set[str]) -> None:
    """The values an answer must name, as opposed to the numbers it computes.

    Counts, hours, and dollars are derived — an agent produces them by
    arithmetic, and they need not appear anywhere. Identifiers are the
    opposite: they can only be copied from something the tools said.
    """

    if isinstance(oracle, str):
        into.add(oracle)
    elif isinstance(oracle, dict):
        for item in oracle.values():
            _identifiers(item, into)
    elif isinstance(oracle, (list, tuple)):
        for item in oracle[:_MAX_FOLLOW]:
            _identifiers(item, into)


def workspace_vocabulary(workspace: Path) -> set[str]:
    """Every string an agent can read off the files it was handed.

    The tool surfaces are not the whole environment. The firm's working
    papers are eighty-five real workbooks sitting in the agent's own
    workspace — fourteen thousand cells that no MCP server serves and no
    SQL query reaches, because a practice keeps them the way a practice
    does, as files. An agent opens them with a shell, which is exactly how
    the person whose job this is would.

    So a workbook's own path, its sheet names, and its cell values are
    reachable in the only sense this module cares about: obtainable
    without guessing. Leaving them out would condemn any task graded on
    the library as unanswerable while an agent answers it.
    """

    found: set[str] = set()
    if not workspace.is_dir():
        return found
    for path in sorted(workspace.rglob("*")):
        if not path.is_file():
            continue
        relative = path.relative_to(workspace)
        found.add(str(relative))
        found.add(path.name)
        found.update(part for part in relative.parts)
        if path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook

                book = load_workbook(path, data_only=True, read_only=True)
            except Exception:
                # A workbook that will not open is a materialization
                # problem, and the gate's job is to report what *is*
                # reachable rather than to fail the build from here.
                continue
            for sheet in book.worksheets:
                found.add(sheet.title)
                for row in sheet.iter_rows(values_only=True):
                    found.update(str(cell).strip() for cell in row if cell is not None)
            book.close()
        elif path.suffix.lower() in (".md", ".txt", ".csv"):
            try:
                text = path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
            for line in text.splitlines():
                found.add(line.strip())
                # Markdown keeps its tables as pipe-delimited rows, so the
                # cell is the unit an oracle would name, not the line.
                if "|" in line:
                    found.update(cell.strip() for cell in line.split("|"))
    found.discard("")
    return found


def unreachable(
    oracle: Any, state_dir: Path, workspace: Path | None = None
) -> list[str]:
    """Identifiers the oracle requires that nothing the agent has ever serves."""

    reachable = served_vocabulary(state_dir)
    if workspace is not None:
        reachable = reachable | workspace_vocabulary(workspace)
    wanted: set[str] = set()
    _identifiers(oracle, wanted)
    return sorted(
        value for value in wanted if _is_identifier(value) and value not in reachable
    )


def solver_columns(solver: Path) -> dict[str, set[str]]:
    """What the reference solver reads, table by column.

    This is documentation, not a gate, and the distinction was expensive
    to learn. ``unreachable`` checks the answer's nouns and deliberately
    ignores derived numbers, which leaves a gap: a task can name only
    reachable engagements while resting on a table nobody can read. The
    regression task counted backward transitions out of ``matter_history``
    when no clio tool returned a status change; Opus 5 scored 0.067 on an
    unobtainable answer.

    Two static gates for it were written and both were vacuous. A table's
    values (``Open``, ``in-progress``, a person's name) all appear
    elsewhere, so "some value is served" passes always; and every surface
    renames columns on the way out (``quantity_seconds`` is served as
    ``quantity``), so "the column name is served" fails always. What
    actually caught the defect was a rollout and a reading of the score.

    So this returns the solver's read set for a human to check against the
    tool list, and the real gate stays where it belongs: any criterion
    below 1.0 is a defect until the transcript says otherwise.
    """

    text = solver.read_text()
    reads: dict[str, set[str]] = {}
    for columns, table in re.findall(
        r"SELECT\s+(.*?)\s+FROM\s+([a-z_][a-z0-9_]*)", text, re.IGNORECASE | re.DOTALL
    ):
        names = {
            part.strip().split()[-1].strip('"')
            for part in columns.split(",")
            if part.strip() and "(" not in part
        }
        reads.setdefault(table.lower(), set()).update(names)
    return reads


def _is_identifier(value: str) -> bool:
    """Something the world minted, as opposed to a word the task defined.

    `tkt-000005` and `Nora Behrens` have to be copied from a tool. `clear`
    and `awaiting_firm_reply` are vocabulary the instruction hands the agent,
    so their absence from the surfaces says nothing.
    """

    value = value.strip()
    if not value or len(value) > 120:
        return False
    if _DATE.fullmatch(value):
        # A calendar date is arithmetic, not vocabulary. Nobody guesses
        # `2026-01-31`; it is what "the end of the month this was sent in"
        # comes to, and the commitment register's due dates land on
        # weekends, on month ends, and past the last day the world ran --
        # so most of them appear in no payload and never could. Treating
        # them as identifiers condemned a task whose every date is derived
        # from a served one by a rule the instruction states in full.
        #
        # This costs the gate nothing it was built for: `tkt-000005` still
        # has digits and is still an identifier, and a date an agent must
        # *copy* is served anyway, so excluding it loses no coverage.
        return False
    if any(character.isdigit() for character in value):
        return True
    # A person or organisation name: capitalised words the agent must match.
    words = value.split()
    return 1 < len(words) < 6 and all(word[:1].isupper() for word in words)
